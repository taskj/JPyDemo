import sys
import json
from openpyxl import Workbook
from openpyxl import load_workbook

'''
Excel文件主要有三个对象
workbook:工作簿，一个excel包含多个sheet
sheet:工作表，一个workbook包含多个sheet
cell:单元格，储存数据对象
'''

#操作现有excel表格类
class loadExcel:

    '''
    示例excel表格列
    编号,姓名,性别,出生年月,政治面貌,籍贯,所在公寓,入学成绩,平均成绩,名次,备注
    '''

    def __init__(self):
        self.excelPath = sys.path[0] + '\students.xlsx' #要读取的xlsx文件路径
        self.wb = load_workbook(self.excelPath) #读取xlsx文件，获得workbook对象
        self.sheets = self.wb.sheetnames #获取所有的工作表
        self.sheet1 = self.wb[self.sheets[0]] #获取Sheet1
        self.rowNum = self.sheet1.max_row  #获取Sheet1工作表行数
        self.colNum = self.sheet1.max_column #获取Sheet1工作表列数
        
    '''
    获取某列的所有值
    @column : 列数
    '''
    def getColVals(self,column):
        colsData = []
        for i in range(2,self.rowNum+1): #样例excel表格内容从第2行开始
            cellVal = self.sheet1.cell(row=i,column=column).value
            colsData.append(cellVal)
        print(colsData)

    '''
    #获取某行的所有值
    @row : 行数
    '''
    def getRowVals(self,row):
        rowsData = []
        for i in range(1,self.colNum):
            cellVal = self.sheet1.cell(row=row,column=i).value
            rowsData.append(cellVal)
        print(rowsData)

    '''
    获取某个单元格的值
    @row : 行数
    @column : 列数
    '''
    def getCellVal(self,row,column):
        cellVal = self.sheet1.cell(row=row,column=column).value
        print(cellVal)
    
    #读取excel中的所有数据并转化成JSON字符串
    def excelToJson(self):
        allData = []
        for i in range(2,self.rowNum+1): #遍历循环sheet1中的所有行数据
            rowData = []
            for j in range(1,self.colNum):
                cellVal = str(self.sheet1.cell(row=i,column=j).value)
                rowData.append(cellVal)
            allData.append(rowData)
        json_str = json.dumps(allData)
        print(json_str)

if __name__ == "__main__":
    loadExcel = loadExcel()
    loadExcel.excelToJson()